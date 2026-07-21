import json
from pathlib import Path
import openpyxl

from .context import CAST_DEFINITION_DIR

DEFAULT_TEMPLATE_DIR = CAST_DEFINITION_DIR
USE_NULL_TARGET = True


def excel_column_name(col_num: int) -> str:
    """Convert a 1-based column number to an Excel column name (e.g., 1 -> A, 27 -> AA)."""
    name = ""
    while col_num:
        col_num, remainder = divmod(col_num - 1, 26)
        name = chr(65 + remainder) + name
    return name


def assess_xlsx_size(xlsx_file: str | Path | None = None) -> tuple[int, int]:
    """Inspects an Excel spreadsheet to auto-detect its actual physical grid boundaries (rows, cols).

    Args:
        xlsx_file: Path to the XLSX workbook. If None or non-existent, defaults to (50, 17).

    Returns:
        tuple[int, int]: (rows, columns)
    """
    if xlsx_file is None or not Path(xlsx_file).exists():
        return (50, 17)

    wb = openpyxl.load_workbook(xlsx_file, read_only=True)
    ws = wb.active
    rows = ws.max_row or 50
    cols = ws.max_column or 17
    wb.close()

    return (rows, cols)


def generate_cast_template(
    source: str = "YYYY",
    target: str = "YYYY",
    rows: int = 50,
    cols: int = 17,
    output_path: str | Path | None = None,
) -> Path:
    """Generate an identity cell mapping template."""
    if rows < 1 or cols < 1:
        raise ValueError("rows and cols must both be >= 1")

    last_col_letter = excel_column_name(cols)
    detected_range = f"A1:{last_col_letter}{rows}"

    cellmap = {}
    for col in range(1, cols + 1):
        col_name = excel_column_name(col)
        for row in range(1, rows + 1):
            cell = f"{col_name}{row}"
            cellmap[cell] = None if USE_NULL_TARGET else cell

    data = {
        "metadata": {
            "source": source,
            "target": target,
            "detected_range": detected_range,
            "description": "Template mapping",
        },
        "cellmap": cellmap,
    }

    if output_path is None:
        output_path = DEFAULT_TEMPLATE_DIR / f"cast_{source}_to_{target}_template.json"
    else:
        output_path = Path(output_path)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with output_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    print(f"Generated cast template ({detected_range}) -> {output_path}")
    return output_path


# Alias for backwards compatibility
generate_blank_template = generate_cast_template
